from flask import Flask, request, send_file, render_template, jsonify
import pandas as pd
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

app = Flask(__name__)
logs = []

def log(msg):
    print(msg)
    logs.append(msg)
    if len(logs) > 1000:
        logs.pop(0)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/logs')
def get_logs():
    return jsonify(logs)

@app.route('/process', methods=['POST'])
def process_excel():
    try:
        logs.clear()
        file = request.files['file']
        df = pd.read_excel(file, header=None)
        total_rows, total_cols = df.shape
        log(f"[开始] 共 {total_rows} 行 x {total_cols} 列")

        image_map = {}
        failed_map = []

        for row in range(total_rows):
            for col in range(1, total_cols):
                cell_value = df.iat[row, col]
                if isinstance(cell_value, str) and cell_value.startswith("http"):
                    url = cell_value.strip()
                    try:
                        headers = {'User-Agent': 'Mozilla/5.0', 'Referer': url}
                        response = requests.get(url, headers=headers, timeout=10)
                        img_data = BytesIO(response.content)
                        with Image.open(img_data) as img:
                            img_io = BytesIO()
                            img.save(img_io, format='PNG')
                            img_io.seek(0)
                            image_map[(row + 1, col + 1)] = img_io
                            log(f"✅ 下载成功: ({row+1},{col+1})")
                    except Exception as e:
                        failed_map.append((row + 1, col + 1))
                        log(f"❌ 下载失败: ({row+1},{col+1}) -> {e}")

        wb = Workbook()
        ws = wb.active
        ws.title = '图片结果'

        for row in range(total_rows):
            name = str(df.iat[row, 0]).strip()
            ws.cell(row=row + 1, column=1, value=name)

        for (r, c), img_io in image_map.items():
            try:
                xl_img = XLImage(img_io)
                xl_img.width = 245
                xl_img.height = 342
                col_letter = chr(64 + c) if c <= 26 else f"{chr(64 + (c - 1) // 26)}{chr(64 + (c - 1) % 26 + 1)}"
                xl_img.anchor = f"{col_letter}{r}"
                ws.row_dimensions[r].height = 342 / 0.75  # ≈ 456
                ws.column_dimensions[col_letter].width = 245 / 7  # ≈ 35
                ws.add_image(xl_img)
            except Exception as e:
                log(f"⚠️ 插图失败: ({r},{c}) -> {e}")
                ws.cell(row=r, column=c, value="插图失败")

        for (r, c) in failed_map:
            ws.cell(row=r, column=c, value="加载失败")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        log("✅ 全部完成！")
        return send_file(output, as_attachment=True, download_name='图片预览结果.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as err:
        log("[服务处理失败] " + str(err))
        return "服务异常：" + str(err), 500

if __name__ == '__main__':
    app.run(debug=True)
